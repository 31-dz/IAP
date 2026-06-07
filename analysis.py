import pandas as pd
import numpy as np
from scipy import stats
import warnings
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings('ignore')

def load_data(filename='lng_flare_data.csv'):
    # Loads flaring data and extracts temporal layers for specialized statistical parsing.
    df = pd.read_csv(filename)
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    df['date'] = df['timestamp'].dt.date
    df['week'] = df['timestamp'].dt.isocalendar().week
    return df

def style_header(ws, row=1):
    # Applies corporate standard styling to worksheet headers.
    for cell in ws[row]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

def descriptive_statistics(df):
    # Calculates comprehensive fundamental and higher-moment descriptive statistics.
    total_rate = df['total_flare_rate_m3_per_hour']
    q1, q3 = total_rate.quantile(0.25), total_rate.quantile(0.75)
    
    metrics = {
        'Count': len(total_rate), 'Mean': total_rate.mean(), 'Median': total_rate.median(),
        'Std Dev': total_rate.std(), 'Min': total_rate.min(), 'Max': total_rate.max(),
        'Range': total_rate.max() - total_rate.min(), 'Q1 (25%)': q1, 'Q3 (75%)': q3,
        'IQR': q3 - q1, 'Skewness': stats.skew(total_rate), 'Kurtosis': stats.kurtosis(total_rate)
    }
    # Add high-resolution percentiles
    for p in [5, 10, 25, 50, 75, 90, 95, 99]:
        metrics[f'{p}th Percentile'] = total_rate.quantile(p / 100.0)
        
    return pd.DataFrame(list(metrics.items()), columns=['Metric', 'Value'])

def normality_tests_by_cause(df):
    # Executes multi-tiered distribution normality verifications across operation modes.
    cause_cols = {
        'normal_operations_m3_per_hour': 'Normal Operations',
        'process_upset_m3_per_hour': 'Process Upset',
        'equipment_maintenance_m3_per_hour': 'Equipment Maintenance',
        'startup_shutdown_m3_per_hour': 'Startup/Shutdown',
        'emergency_relief_m3_per_hour': 'Emergency Relief',
        'compressor_trip_m3_per_hour': 'Compressor Trip',
        'instrument_failure_m3_per_hour': 'Instrument Failure'
    }
    results = []
    for col, name in cause_cols.items():
        active_data = df.loc[df[col] > 0, col]
        
        if len(active_data) < 3:
            results.append({
                'Cause': name, 'Sample Size': len(active_data), 'Mean (m³/hr)': 0, 'Std Dev': 0,
                'Shapiro-Wilk Statistic': 'N/A', 'Shapiro-Wilk P-value': 'N/A', 'Shapiro Result': 'Insufficient Data',
                'KS Statistic': 'N/A', 'KS P-value': 'N/A', 'KS Result': 'Insufficient Data',
                'Anderson Statistic': 'N/A', 'Interpretation': 'Not enough data for testing'
            })
            continue
        
        # Sub-sample if sequence limit compromises Shapiro-Wilk validity constraints (N > 5000)
        sample = active_data.sample(n=min(5000, len(active_data)), random_state=42)
        shapiro_stat, shapiro_p = stats.shapiro(sample)
        
        ks_stat, ks_p = stats.kstest(active_data, 'norm', args=(active_data.mean(), active_data.std()))
        anderson_stat = stats.anderson(active_data, dist='norm').statistic
        
        shapiro_normal = shapiro_p > 0.05
        ks_normal = ks_p > 0.05
        
        if shapiro_normal and ks_normal:
            interpretation = 'Data appears normally distributed'
        elif not shapiro_normal and not ks_normal:
            interpretation = 'Data is NOT normally distributed'
        else:
            interpretation = 'Mixed results - likely not normal'
            
        results.append({
            'Cause': name, 'Sample Size': len(active_data), 'Mean (m³/hr)': round(active_data.mean(), 2),
            'Std Dev': round(active_data.std(), 2), 'Shapiro-Wilk Statistic': round(shapiro_stat, 6),
            'Shapiro-Wilk P-value': f"{shapiro_p:.6e}" if shapiro_p < 0.001 else round(shapiro_p, 6),
            'Shapiro Result': 'Normal' if shapiro_normal else 'Not Normal', 'KS Statistic': round(ks_stat, 6),
            'KS P-value': f"{ks_p:.6e}" if ks_p < 0.001 else round(ks_p, 6), 'KS Result': 'Normal' if ks_normal else 'Not Normal',
            'Anderson Statistic': round(anderson_stat, 6), 'Interpretation': interpretation
        })
    return pd.DataFrame(results)

def outlier_analysis(df):
    # Evaluates mathematical anomalies via IQR, Z-Score, and Modified Z-Score methods.
    total_rate = df['total_flare_rate_m3_per_hour']
    
    # Method 1: Interquartile Range
    q1, q3 = total_rate.quantile(0.25), total_rate.quantile(0.75)
    iqr = q3 - q1
    lower_bound, upper_bound = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    iqr_outliers = df[(total_rate < lower_bound) | (total_rate > upper_bound)]
    
    # Method 2: Standard Z-Score
    z_scores = np.abs(stats.zscore(total_rate))
    z_outliers = df[z_scores > 3]
    
    # Method 3: Modified Z-Score using Median Absolute Deviation (MAD)
    median = total_rate.median()
    mad = np.median(np.abs(total_rate - median))
    mod_z = 0.6745 * (total_rate - median) / mad if mad != 0 else np.zeros(len(total_rate))
    mod_z_outliers = df[np.abs(mod_z) > 3.5]
    
    summary = pd.DataFrame({
        'Method': ['IQR (1.5× IQR)', 'Z-Score (|z| > 3)', 'Modified Z-Score (MAD, |z| > 3.5)'],
        'Outliers Found': [len(iqr_outliers), len(z_outliers), len(mod_z_outliers)],
        'Percentage': [f"{len(x)/len(df)*100:.2f}%" for x in [iqr_outliers, z_outliers, mod_z_outliers]],
        'Lower Bound': [f"{lower_bound:.2f}", 'N/A', 'N/A'],
        'Upper Bound': [f"{upper_bound:.2f}", 'N/A', 'N/A']
    })
    
    top_outliers = pd.DataFrame(columns=['Timestamp', 'Flare Rate (m³/hr)', 'Dominant Cause', 'Severity'])
    if not iqr_outliers.empty:
        top_outliers = iqr_outliers.nlargest(10, 'total_flare_rate_m3_per_hour')[
            ['timestamp', 'total_flare_rate_m3_per_hour', 'dominant_cause', 'severity']
        ].copy()
        top_outliers.columns = ['Timestamp', 'Flare Rate (m³/hr)', 'Dominant Cause', 'Severity']
        
    return summary, top_outliers

def cause_correlation_analysis(df):
    # Unveils structural dependencies between different flare generation root causes.
    cause_cols = [
        'normal_operations_m3_per_hour', 'process_upset_m3_per_hour', 'equipment_maintenance_m3_per_hour',
        'startup_shutdown_m3_per_hour', 'emergency_relief_m3_per_hour', 'compressor_trip_m3_per_hour',
        'instrument_failure_m3_per_hour'
    ]
    short_names = ['Normal Ops', 'Process Upset', 'Equip Maint', 'Startup/SD', 'Emergency', 'Compressor', 'Instrument']
    
    corr_matrix = df[cause_cols].corr().round(4)
    corr_matrix.columns = short_names
    corr_matrix.index = short_names
    
    strong_corr = []
    raw_corr = df[cause_cols].corr()
    for i in range(len(cause_cols)):
        for j in range(i + 1, len(cause_cols)):
            val = raw_corr.iloc[i, j]
            if abs(val) > 0.05:
                strong_corr.append({'Cause 1': short_names[i], 'Cause 2': short_names[j], 'Correlation': round(val, 4)})
                
    strong_corr_df = pd.DataFrame(strong_corr).sort_values('Correlation', key=abs, ascending=False) if strong_corr else pd.DataFrame(columns=['Cause 1', 'Cause 2', 'Correlation'])
    return corr_matrix, strong_corr_df

def temporal_analysis(df):
    # Calculates cyclical aggregate characteristics across multiple granular tracking intervals.
    hourly = df.groupby('hour')['total_flare_rate_m3_per_hour'].agg(['mean', 'std', 'min', 'max', 'count']).round(2)
    hourly.columns = ['Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Count']
    
    dow_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    daily = df.groupby('day_of_week')['total_flare_rate_m3_per_hour'].agg(['mean', 'std', 'min', 'max', 'count']).reindex(dow_order).round(2)
    daily.columns = ['Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Count']
    
    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    monthly = df.groupby('month')['total_flare_rate_m3_per_hour'].agg(['mean', 'sum', 'std', 'min', 'max', 'count']).reindex(month_order).round(2)
    monthly.columns = ['Mean (m³/hr)', 'Total (m³)', 'Std Dev', 'Min', 'Max', 'Count']
    
    weekly = df.groupby('week')['total_flare_rate_m3_per_hour'].sum().round(2).reset_index()
    weekly.columns = ['Week', 'Total Flare (m³)']
    
    return hourly, daily, monthly, weekly

def severity_analysis(df):
    # Profiles volumetric impact metrics filtered strictly by risk-classification levels.
    stats_df = df.groupby('severity')['total_flare_rate_m3_per_hour'].agg(['count', 'mean', 'std', 'min', 'max', 'sum']).round(2)
    stats_df.columns = ['Count', 'Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Total (m³)']
    stats_df['Percentage'] = (stats_df['Count'] / len(df) * 100).round(2)
    return stats_df[['Count', 'Percentage', 'Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Total (m³)']]

def cause_specific_statistics(df):
    # Extracts structural asset operational metrics unique to each source variant.
    cause_cols = {
        'normal_operations_m3_per_hour': 'Normal Operations',
        'process_upset_m3_per_hour': 'Process Upset',
        'equipment_maintenance_m3_per_hour': 'Equipment Maintenance',
        'startup_shutdown_m3_per_hour': 'Startup/Shutdown',
        'emergency_relief_m3_per_hour': 'Emergency Relief',
        'compressor_trip_m3_per_hour': 'Compressor Trip',
        'instrument_failure_m3_per_hour': 'Instrument Failure'
    }
    results = []
    for col, name in cause_cols.items():
        data = df[col]
        active = data[data > 0]
        results.append({
            'Cause': name, 'Total Volume (m³)': round(data.sum(), 2), 'Active Hours': len(active),
            'Active %': f"{len(active)/len(df)*100:.2f}%",
            'Mean Active (m³/hr)': round(active.mean(), 2) if not active.empty else 0,
            'Std Dev Active': round(active.std(), 2) if len(active) > 1 else 0,
            'Max Rate (m³/hr)': round(active.max(), 2) if not active.empty else 0
        })
    return pd.DataFrame(results)

def dominant_cause_analysis(df):
    # Quantifies allocation density distribution patterns based on highest single-point contributors.
    counts = df['dominant_cause'].value_counts().reset_index()
    counts.columns = ['Cause', 'Frequency']
    counts['Percentage'] = (counts['Frequency'] / len(df) * 100).round(2)
    
    cause_mapping = {
        'normal_operations': 'Normal Operations', 'process_upset': 'Process Upset',
        'equipment_maintenance': 'Equipment Maintenance', 'startup_shutdown': 'Startup/Shutdown',
        'emergency_relief': 'Emergency Relief', 'compressor_trip': 'Compressor Trip',
        'instrument_failure': 'Instrument Failure'
    }
    counts['Cause'] = counts['Cause'].map(cause_mapping)
    return counts

def trend_analysis(df):
    # Calculates macro trajectory parameters alongside historical baseline tracking.
    daily = df.groupby('date')['total_flare_rate_m3_per_hour'].sum().reset_index()
    daily['day_number'] = range(len(daily))
    
    slope, _, r_val, p_val, std_err = stats.linregress(daily['day_number'], daily['total_flare_rate_m3_per_hour'])
    
    if p_val < 0.05:
        trend_desc = 'Significant increasing trend' if slope > 0 else 'Significant decreasing trend'
    else:
        trend_desc = 'No significant trend'
        
    trend_summary = pd.DataFrame({
        'Metric': ['Slope (m³/day per day)', 'R-squared', 'P-value', 'Std Error', 'Interpretation'],
        'Value': [round(slope, 4), round(r_val**2, 4), f"{p_val:.6e}" if p_val < 0.001 else round(p_val, 6), round(std_err, 4), trend_desc]
    })
    
    df_sorted = df.sort_values('timestamp')
    ma_summary = pd.DataFrame({
        'Moving Average': ['24-hour MA', '7-day MA (168h)'],
        'Min (m³/hr)': [round(df_sorted['total_flare_rate_m3_per_hour'].rolling(w).mean().min(), 2) for w in [24, 168]],
        'Max (m³/hr)': [round(df_sorted['total_flare_rate_m3_per_hour'].rolling(w).mean().max(), 2) for w in [24, 168]],
        'Mean (m³/hr)': [round(df_sorted['total_flare_rate_m3_per_hour'].rolling(w).mean().mean(), 2) for w in [24, 168]]
    })
    return trend_summary, ma_summary

def run_weekly_year_span_sarima(df, output_csv='normal_ops_forecast_2026.csv'):
    # Fits an autoregressive SARIMA state space structure for predicting future baseline configurations.
    from statsmodels.tsa.statespace.sarimax import SARIMAX
    
    # Memory efficient slicing prevents mutating global dataframe indices
    ts_data = df[['timestamp', 'normal_operations_m3_per_hour']].copy().set_index('timestamp')
    weekly_series = ts_data['normal_operations_m3_per_hour'].resample('W').mean()
    
    model = SARIMAX(weekly_series, order=(1, 1, 1), seasonal_order=(0, 1, 0, 52),
                    enforce_stationarity=False, enforce_invertibility=False)
    results = model.fit(disp=False)
    
    forecast_steps = 52
    forecast_res = results.get_forecast(steps=forecast_steps)
    forecast_index = pd.date_range(start='2026-01-04', periods=forecast_steps, freq='W')
    raw_forecast = forecast_res.predicted_mean.values
    
    # Boundary smoothing alignment
    initial_gap = weekly_series.values[-1] - raw_forecast[0]
    smoothed_forecast = np.copy(raw_forecast)
    for i in range(min(4, forecast_steps)):
        weight = 1.0 - ((i + 1) / 5)
        smoothed_forecast[i] += (initial_gap * weight)
        
    forecast_df = pd.DataFrame({
        'date': forecast_index.strftime('%Y-%m-%d'),
        'forecasted_normal_ops_m3_per_hour': np.clip(smoothed_forecast, a_min=0, a_max=None)
    })
    
    forecast_df.to_csv(output_csv, index=False)
    return forecast_df

def main():
    # Ask the user to input the path to the input data file instead of searching for it directly
    filename = input("Please enter the path to the LNG flare data CSV file (default: lng_flare_data.csv): ")
    if not filename.strip():
        filename = 'lng_flare_data.csv'
        
    df = load_data(filename)
    
    # Ask the user to input the path for the forecasted output file
    output_forecast_csv = input("Please enter the path/filename to save the forecasted SARIMA CSV file (default: normal_ops_forecast_2026.csv): ")
    if not output_forecast_csv.strip():
        output_forecast_csv = 'normal_ops_forecast_2026.csv'
        
    # Ask the user to input the path for the final Excel statistical report
    output_excel = input("Please enter the path/filename to save the statistical analysis Excel report (default: lng_flare_statistical_analysis.xlsx): ")
    if not output_excel.strip():
        output_excel = 'lng_flare_statistical_analysis.xlsx'
        
    forecast_df = run_weekly_year_span_sarima(df, output_csv=output_forecast_csv)
    output_file = output_excel
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: Executive Overview
        pd.DataFrame({
            'Metric': ['Total Sample Records', 'Historical Date Range', 'Total Flaired Mass (m³)', 'Average Rate (m³/hr)', 'Peak Rate (m³/hr)'],
            'Value': [len(df), f"{df['timestamp'].min()} to {df['timestamp'].max()}", df['total_flare_rate_m3_per_hour'].sum(), df['total_flare_rate_m3_per_hour'].mean(), df['total_flare_rate_m3_per_hour'].max()]
        }).to_excel(writer, sheet_name='Summary', index=False)
        
        # Core Statistical Layers
        descriptive_statistics(df).to_excel(writer, sheet_name='Descriptive Stats', index=False)
        normality_tests_by_cause(df).to_excel(writer, sheet_name='Normality Tests', index=False)
        
        # Multi-Table Sheet Formats
        outlier_summary, top_outliers = outlier_analysis(df)
        outlier_summary.to_excel(writer, sheet_name='Outliers', index=False, startrow=0)
        if not top_outliers.empty:
            top_outliers.to_excel(writer, sheet_name='Outliers', index=False, startrow=len(outlier_summary) + 3)
            
        corr_matrix, strong_corr = cause_correlation_analysis(df)
        corr_matrix.to_excel(writer, sheet_name='Correlations', startrow=0)
        if not strong_corr.empty:
            strong_corr.to_excel(writer, sheet_name='Correlations', index=False, startrow=len(corr_matrix) + 3)
            
        # Chronological Breakdowns
        hourly, daily, monthly, weekly = temporal_analysis(df)
        hourly.to_excel(writer, sheet_name='Temporal-Hourly')
        daily.to_excel(writer, sheet_name='Temporal-Daily')
        monthly.to_excel(writer, sheet_name='Temporal-Monthly')
        weekly.to_excel(writer, sheet_name='Temporal-Weekly', index=False)
        
        # Operational Categorization Blocks
        severity_analysis(df).to_excel(writer, sheet_name='Severity Analysis')
        cause_specific_statistics(df).to_excel(writer, sheet_name='Cause Statistics', index=False)
        dominant_cause_analysis(df).to_excel(writer, sheet_name='Dominant Causes', index=False)
        
        # Macro Trends & Forecasting Outputs
        trend_sum, ma_sum = trend_analysis(df)
        trend_sum.to_excel(writer, sheet_name='Trend Analysis', index=False, startrow=0)
        ma_sum.to_excel(writer, sheet_name='Trend Analysis', index=False, startrow=len(trend_sum) + 3)
        
        forecast_df.to_excel(writer, sheet_name='SARIMA 2026 Forecast', index=False)
        
        # Global uniform structural formatting sweep
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            style_header(ws)
            for column in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_len + 3, 50)
                
    print(f"✓ Full report analysis compilation completed successfully. Output files saved to: {output_forecast_csv} and {output_excel}")

if __name__ == "__main__":
    main()
